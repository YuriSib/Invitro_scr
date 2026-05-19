from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


class ExcelTable:
    def __init__(self, file_path: str, sheet_name: str | None = None):
        self.file_path = file_path

        self._insert_doc(self.file_path)

        self.wb = load_workbook(file_path)
        self.ws = (
            self.wb[sheet_name]
            if sheet_name
            else self.wb.active
        )

        self.default_font = Font(name="Calibri", size=10)
        self.header_font = Font(name="Calibri", size=10, bold=True)

        self._set_columns()

    def _insert_doc(self, file_path):
        wb = Workbook()
        wb.save(file_path)
        wb.close()

    def save(self):
        self.wb.save(self.file_path)

    def save_and_close(self):
        self.wb.save(self.file_path)
        self.wb.close()

    def _set_columns(self):
        # Записываем заголовки в первую строку
        headers = ["Название анализа", "Стоимость, руб."]
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header

            # оформление
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")

        # Устанавливаем ширину колонок
        for col, width in [(1, 120), (2, 18)]:
            col_letter = get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = width

    def add_category_row(self, category_name: str):
        last_row = self.ws.max_row + 1

        self.ws.merge_cells(start_row=last_row, start_column=1,
                       end_row=last_row, end_column=2)

        cell = self.ws.cell(row=last_row, column=1)
        cell.value = category_name

        cell.alignment = Alignment(horizontal="center")
        cell.font = self.header_font


    def add_service(self, service: dict) -> str:
        # Проверка структуры
        required_keys = {"title", "price"}

        if not isinstance(service, dict):
            return "product должен быть dict"

        if not required_keys.issubset(service):
            return "product должен содержать ключи: title, price"

        title = service["title"]
        price = service["price"]

        if not isinstance(title, str):
            return  "title должен быть строкой"

        if not isinstance(price, (int, float)):
            return  "price должен быть числом"

        self.ws.append([title, price])

        last_row = self.ws.max_row

        for col in range(1, 3):
            cell = self.ws.cell(row=last_row, column=col)
            cell.font = self.default_font

        return "OK"

    def add_services(self, services: list[dict]) -> str:
        for product in services:
            result = self.add_service(product)
            if result != "OK":
                return f"Были переданы невалидные данные. {result}"

        return "OK"

